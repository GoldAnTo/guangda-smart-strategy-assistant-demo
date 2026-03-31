import openpyxl, json

wb = openpyxl.load_workbook(r'E:\code\guangda-smart-strategy-assistant-demo\strategy_market_new.xlsx', data_only=True)

def pct(v):
    if v is None: return None
    s = str(v).replace('%','').strip()
    try: return float(s)
    except: return None

def parse_strategy_meta(row):
    # 0:序号 1:策略id 2:策略名称 3:来源 4:简称 5:结构 6:星级
    # 7:维护人 8:开始日期 9:策略逻辑 10:出品机构 11:适合客户
    # 12:业务文件 13:上线日 14:业务标签
    # 15:年胜率% 16:月胜率% 17:年化收益% 18:年化波动% 19:最大回撤% 20:夏普比率 21:基准名称
    sid = str(row[1]).strip() if row[1] else ''
    if not sid: return None
    name = str(row[2]).strip() if row[2] else ''
    source = str(row[3]).strip() if row[3] else ''
    structure = str(row[5]).strip() if row[5] else ''
    stars = int(row[6]) if row[6] else 3
    owner = str(row[7]).strip() if row[7] else ''
    start_date = str(row[8]).strip() if row[8] else ''
    logic = str(row[9]).strip() if row[9] else ''
    suitable = str(row[11]).strip() if row[11] else ''
    tags_str = str(row[14]).strip() if row[14] else ''
    tags = [t.strip() for t in tags_str.split(',') if t.strip()]
    annual_return = pct(row[17])
    volatility = pct(row[18])
    max_drawdown = pct(row[19])
    sharpe_raw = row[20]
    sharpe = float(sharpe_raw) if sharpe_raw and str(sharpe_raw).strip() else None
    if sharpe and (sharpe < -50 or sharpe > 50): sharpe = None
    annual_win = pct(row[15])
    monthly_win = pct(row[16])
    benchmark = str(row[21]).strip() if row[21] else ''

    cat_map = {'转债':'转债','股票':'股票','ETF':'ETF','股指期货':'股指期货','商品':'商品','混合型':'混合型','股基FOF':'股基FOF'}
    category = cat_map.get(source, source)
    nav_cat = source or '其他'
    risk_map = {'转债':'R3','混合型':'R3','商品':'R4','股指期货':'R4','股基FOF':'R4','ETF':'R4','股票':'R5'}
    risk = risk_map.get(source, 'R3')
    risk_disp = {'R3':'中等风险','R4':'中高风险','R5':'高风险'}
    horizon_map = {'雪球':'medium_term','全天候':'long_term','绝对收益':'medium_term'}
    horizon = horizon_map.get(name[:3], 'medium_term') if name else 'medium_term'
    horizon_disp = {'short_term':'短期交易','medium_term':'中期配置','long_term':'长期配置'}
    liq = 'high' if 'ETF' in source else 'medium'
    liq_disp = '高流动性' if liq == 'high' else '中等流动性'

    return {
        'id': f'GS-{sid}',
        'seed': int(sid),
        'name': name,
        'navCategory': nav_cat,
        'category': category,
        'productType': category.lower().replace(' ', '_'),
        'riskLevel': risk,
        'riskLevelDisplay': risk_disp.get(risk,'中等风险'),
        'investmentHorizon': horizon,
        'investmentHorizonDisplay': horizon_disp.get(horizon,'中期配置'),
        'liquidity': liq,
        'liquidityDisplay': liq_disp,
        'returnExpectation': 'balanced',
        'returnExpectationDisplay': '追求平衡收益',
        'volatility': 'medium',
        'volatilityDisplay': '中等波动',
        'structure': structure,
        'owner': owner,
        'startDate': start_date,
        'logicSummary': logic,
        'benchmarkName': benchmark,
        'positioning': suitable,
        'tags': tags,
        'outlookStars': stars,
        'annualReturn': annual_return if annual_return is not None else 0.0,
        'winRate': annual_win if annual_win is not None else 0.0,
        'maxDrawdown': max_drawdown,
        'volatilityValue': volatility,
        'sharpe': sharpe,
        'monthlyWinRate': monthly_win,
        'suitableFor': [t for t in suitable.split('、') if t] if suitable else [],
        'notSuitableFor': [],
        'matchKeywords': tags,
        'standardRiskNotice': '历史收益不代表未来表现，投资有风险。'
    }

# ── 策略元数据 ──────────────────────────────────────────
ws_meta = wb[wb.sheetnames[0]]
strategies = []
for row in ws_meta.iter_rows(min_row=2, values_only=True):
    s = parse_strategy_meta(row)
    if s: strategies.append(s)

strategies.sort(key=lambda x: x['seed'])
print(f'策略元数据: {len(strategies)} 条')
for s in strategies:
    print(f"  {s['id']} | {s['name']:<12} | {s['navCategory']:<8} | 年化{s['annualReturn']:>7}% | 胜率{s['winRate']:>6}% | 夏普{s['sharpe']} | 回撤{s['maxDrawdown']} | {s['benchmarkName']}")

# ── 历史净值 (成立以来) ────────────────────────────────
ws_ts = wb[wb.sheetnames[5]]
ts_data = {}  # sid(str) -> [{date, ret, benchmark}]
for row in ws_ts.iter_rows(min_row=2, values_only=True):
    sid = str(row[0]).strip() if row[0] else ''
    date = str(int(row[1])) if row[1] else ''  # date as int like 20260228
    if not sid or not date: continue
    if sid not in ts_data: ts_data[sid] = []
    # 计算每日收益率（从净值变化）
    prev = ts_data[sid][-1] if ts_data[sid] else None
    nav_str = str(row[4]).replace('%','') if row[4] else None
    bm_str = str(row[5]).replace('%','') if row[5] else None
    try:
        nav_val = float(nav_str) if nav_str else 0.0
        bm_val = float(bm_str) if bm_str else 0.0
    except:
        nav_val, bm_val = 0.0, 0.0
    ts_data[sid].append({'date': date, 'ret': nav_val, 'benchmark': bm_val})

# 只保留有数据的策略
ts_clean = {}
for s in strategies:
    sid = str(s['seed'])
    if sid in ts_data and ts_data[sid]:
        # 排序
        ts_clean[sid] = sorted(ts_data[sid], key=lambda x: x['date'])
    else:
        ts_clean[sid] = []

print(f'\n净值数据: {sum(1 for v in ts_clean.values() if v)} 条策略有NAV记录')
for sid, pts in sorted(ts_clean.items()):
    if pts:
        print(f"  seed={sid}: {len(pts)} 条, {pts[0]['date']} ~ {pts[-1]['date']}")

# ── 写入文件 ───────────────────────────────────────────
with open(r'E:\code\guangda-smart-strategy-assistant-demo\server\src\data\strategies-enhanced.json','w',encoding='utf-8') as f:
    json.dump({'version':'2.0','exported_at':'2026-03-31','strategies':strategies}, f, ensure_ascii=False, indent=2)
print('\n✅ strategies-enhanced.json 已写入')

with open(r'E:\code\guangda-smart-strategy-assistant-demo\server\src\data\strategies-timeseries.json','w',encoding='utf-8') as f:
    json.dump(ts_clean, f, ensure_ascii=False)
print('✅ strategies-timeseries.json 已写入')
